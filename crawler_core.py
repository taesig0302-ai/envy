# -*- coding: utf-8 -*-
import os, re, json, time, random, html, hashlib
from typing import List, Dict, Tuple
from urllib.parse import urlparse
import openpyxl, requests
from requests.adapters import HTTPAdapter, Retry
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# ──────────────────────────────────────────────────────────────
# 설정
# ──────────────────────────────────────────────────────────────
BASE_DIR = r'C:\singlefiless'
IMG_ROOT = os.path.join(BASE_DIR, 'images')
EXCEL_PATH = os.path.join(BASE_DIR, '네이버_상품정보_수집_확장.xlsx')
os.makedirs(IMG_ROOT, exist_ok=True)
os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)

DOWNLOAD_IMAGES = True
WAIT_SEC = 12
PAGELOAD_TIMEOUT = 25
LOAD_RETRIES = 2
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36"

CHROMEDRIVER_PATH = r"C:/PATH/to/chromedriver.exe"
MARKETS = [
    "https://smartstore.naver.com/nutri_health/category/a932ac29907c43a9bda9c1e60ab0e244?st=TOTALSALE&dt=IMAGE&page=1&size=60",
]
AUTOSAVE_EVERY = 10

# ──────────────────────────────────────────────────────────────
# 브라우저
# ──────────────────────────────────────────────────────────────
def make_driver() -> Tuple[webdriver.Chrome, WebDriverWait]:
    opts = Options()
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--lang=ko-KR")
    opts.add_argument("start-maximized")
    opts.add_argument(f"user-agent={UA}")
    service = Service(CHROMEDRIVER_PATH)
    d = webdriver.Chrome(service=service, options=opts)
    d.set_page_load_timeout(PAGELOAD_TIMEOUT)
    d.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"}
    )
    return d, WebDriverWait(d, WAIT_SEC)

# ──────────────────────────────────────────────────────────────
# 엑셀
# ──────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
PLACEHOLDER_NAME = "_init"
wb.active.title = PLACEHOLDER_NAME

def ensure_ws(name: str):
    safe = re.sub(r'[\/:*?"<>|]', "_", name)[:31]
    ws = wb.create_sheet(safe)
    ws.append([
        "번호","상품번호","상품명","정가","할인가","배송비","리뷰수","평점",
        "상세페이지 URL","원본상품ID","이미지폴더","첫이미지","저장파일목록","이미지수"
    ])
    if PLACEHOLDER_NAME in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb[PLACEHOLDER_NAME]
    return ws

def safe_save_workbook(wb, path: str):
    tmp = path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, path)
    if os.path.exists(tmp):
        try: os.remove(tmp)
        except: pass

# ──────────────────────────────────────────────────────────────
# 유틸
# ──────────────────────────────────────────────────────────────
IMG_EXT_RE = re.compile(r'\.(?:jpg|jpeg|png|gif|bmp)(?:\?|$)', re.IGNORECASE)
PROD_ID_RE = re.compile(r"/products/(\d+)")

def hash_md5(s: str): return hashlib.md5(s.encode("utf-8")).hexdigest()

def extract_product_id(href: str, kv: Dict[str, str]) -> str:
    for k in ("chnl_prod_no","chnl_prod_id","productNo","chnlPordNo"):
        v = kv.get(k)
        if v: return str(v).strip()
    m = PROD_ID_RE.search(href or "")
    return m.group(1) if m else hash_md5(href)[:10]

def market_id_from_url(url: str) -> str:
    path = urlparse(url).path.strip('/')
    return re.sub(r'[^0-9a-zA-Z_ㄱ-힣-]', '_', path.split('/')[0]) or "market"

# ──────────────────────────────────────────────────────────────
# 상품 목록 로딩
# ──────────────────────────────────────────────────────────────
def ensure_products_loaded(driver: webdriver.Chrome, wait: WebDriverWait, retries: int = LOAD_RETRIES) -> bool:
    sel = (By.CSS_SELECTOR, 'a.linkAnchor[data-shp-contents-dtl]')
    for _ in range(retries):
        try:
            wait.until(EC.presence_of_all_elements_located(sel))
            return True
        except TimeoutException:
            try: driver.refresh()
            except: pass
            time.sleep(1)
    return False

# ──────────────────────────────────────────────────────────────
# 이미지
# ──────────────────────────────────────────────────────────────
SESSION = requests.Session()
SESSION.mount("https://", HTTPAdapter(max_retries=Retry(total=3, backoff_factor=0.3)))
SESSION.headers.update({"User-Agent": UA})

def save_image_as_seq(url: str, folder: str, prod_no: str, seq: int, referer: str) -> str:
    try:
        os.makedirs(folder, exist_ok=True)
        r = SESSION.get(url, headers={"Referer": referer}, timeout=10)
        if r.status_code != 200 or len(r.content) < 1024:
            return ""
        ext = re.search(r'\.(jpg|jpeg|png|gif|bmp)', url.split("?")[0], re.I)
        ext = "." + (ext.group(1).lower() if ext else "jpg")
        path = os.path.join(folder, f"{prod_no}_{seq:02d}{ext}")
        with open(path, "wb") as f:
            f.write(r.content)
        return path
    except Exception:
        return ""

def _push_candidates_from_img(el, seen: set, out: list):
    cand = [
        el.get_attribute("src"),
        el.get_attribute("data-src"),
        el.get_attribute("data-lazy-src"),
        el.get_attribute("currentSrc"),
    ]
    ss = el.get_attribute("srcset") or ""
    if ss:
        parts = [p.strip().split(" ")[0] for p in ss.split(",") if p.strip()]
        cand.extend(parts)
    for u in cand:
        if not u or not u.startswith("http"): 
            continue
        # 확장자 체크 (쿼리 이전 기준)
        if not IMG_EXT_RE.search(u.split("?")[0]): 
            continue
        key = u.lower()  # ❗중복은 '완전 동일 URL' 기준 (사이즈/프레임 분리)
        if key in seen: 
            continue
        seen.add(key)
        out.append(u)

def _push_background_images(scope_el, seen: set, out: list):
    # style="background-image:url(...)" 형태도 수집
    els = scope_el.find_elements(By.CSS_SELECTOR, "[style*='background-image']")
    for e in els:
        style = e.get_attribute("style") or ""
        # url("...") 또는 url('...') 또는 url(...)
        m = re.search(r"url\((['\"]?)(.+?)\1\)", style)
        if not m: 
            continue
        u = m.group(2)
        if not u.startswith("http"): 
            continue
        if not IMG_EXT_RE.search(u.split("?")[0]): 
            continue
        key = u.lower()
        if key in seen: 
            continue
        seen.add(key)
        out.append(u)

def collect_card_img_urls(card_el) -> List[str]:
    """
    상품 카드 내 모든 이미지 URL 수집:
      1) 카드 컨테이너 내부 <img>
      2) 같은 카드의 <li> 내 swiper-wrapper 전체 <img>
      3) background-image:url(...) 형태
      4) src/data-src/data-lazy-src/currentSrc/srcset 모두
    """
    urls, seen = [], set()

    # 1) 카드 컨테이너
    try:
        container = card_el.find_element(By.CSS_SELECTOR, '.I3B6dXSHqa, .zslOZxOl9K')
    except Exception:
        container = card_el

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", container)
        time.sleep(random.uniform(0.15, 0.45))
    except Exception:
        pass

    for img in container.find_elements(By.CSS_SELECTOR, "img"):
        _push_candidates_from_img(img, seen, urls)
    _push_background_images(container, seen, urls)

    # 2) 같은 카드의 <li> 기준 swiper-wrapper 전체 이미지 (보이지 않는 슬라이드 포함)
    try:
        li = card_el.find_element(By.XPATH, "./ancestor::li[1]")
        for img in li.find_elements(By.CSS_SELECTOR, ".swiper-wrapper img"):
            _push_candidates_from_img(img, seen, urls)
        _push_background_images(li, seen, urls)
    except Exception:
        pass

    return urls

# ──────────────────────────────────────────────────────────────
# 크롤링 (세이프가드 저장 포함)
# ──────────────────────────────────────────────────────────────
def collect_one_market(driver: webdriver.Chrome, wait: WebDriverWait, user_url: str):
    market_id = market_id_from_url(user_url)
    ws = ensure_ws(market_id)
    prod_map, last_no = load_index(market_id)

    print(f"\n=== {market_id} ===")
    driver.get(user_url)
    input("✅ 캡차 풀었으면 엔터… ")
    ok = ensure_products_loaded(driver, wait)
    if not ok:
        print("⚠️ 상품 목록 로딩 실패")
        return

    cards = driver.find_elements(By.CSS_SELECTOR, 'a.linkAnchor[data-shp-contents-dtl]')
    print("상품 수:", len(cards))

    row, autosave_counter = 1, 0

    try:
        for tag in cards:
            try:
                detail_json = tag.get_attribute("data-shp-contents-dtl") or ""
                data = json.loads(html.unescape(detail_json)) if detail_json else []
                kv = {d.get("key"): d.get("value") for d in data if isinstance(d, dict)}
                title = kv.get("chnl_prod_nm", "")
                price = kv.get("price", "")
                href = tag.get_attribute("href") or ""
                if not href:
                    continue

                product_id = extract_product_id(href, kv)
                product_no, last_no = assign_product_no(product_id, prod_map, last_no)

                # li 기준 부가 필드
                sale_price = shipping_fee = review_count = rating = ""
                try:
                    li = tag.find_element(By.XPATH, "./ancestor::li[1]")
                except:
                    li = None

                if li:
                    # 할인가
                    try:
                        sale_el = li.find_element(By.CSS_SELECTOR, "span.zIK_uvWc6D")
                        sale_price = re.sub(r"[^\d]", "", sale_el.text)
                    except:
                        try:
                            sale_el = li.find_element(By.XPATH, './/*[contains(text(),"원") and not(contains(text(),"배송"))]')
                            sale_price = re.sub(r"[^\d]", "", sale_el.text)
                        except:
                            pass

                    # 배송비
                    try:
                        if li.find_elements(By.XPATH, './/*[contains(text(),"무료배송")]'):
                            shipping_fee = "0"
                        else:
                            fee_nodes = li.find_elements(By.CSS_SELECTOR, "div.UVrxHKBc0E")
                            got_fee = False
                            for node in fee_nodes:
                                txt = (node.text or "").strip()
                                digits = re.sub(r"[^\d]", "", txt)
                                if digits:
                                    shipping_fee = digits
                                    got_fee = True
                                    break
                            if not got_fee:
                                try:
                                    ship_el = li.find_element(By.XPATH, './/*[contains(text(),"배송")]')
                                    t = ship_el.text
                                    m = re.search(r"(\d[\d,]*)\s*원", t)
                                    shipping_fee = m.group(1).replace(",", "") if m else re.sub(r"[^\d]", "", t)
                                except:
                                    shipping_fee = ""
                    except:
                        shipping_fee = ""

                    # 리뷰/평점
                    try:
                        rv = li.find_element(By.XPATH, './/*[contains(text(),"리뷰") or contains(@class,"GF9")]')
                        review_count = re.sub(r"[^\d]", "", rv.text)
                    except: pass
                    try:
                        rt = li.find_element(By.XPATH, './/*[contains(text(),"평점") or contains(text(),"★")]')
                        m = re.search(r"[\d\.]+", rt.text)
                        rating = m.group(0) if m else ""
                    except: pass

                # 이미지 — 여러 장 모두 저장
                img_folder = os.path.join(IMG_ROOT, market_id, product_no)
                urls = collect_card_img_urls(tag)
                saved_rel = []
                if DOWNLOAD_IMAGES and urls:
                    for i, u in enumerate(urls, start=1):
                        p = save_image_as_seq(u, img_folder, product_no, i, referer=user_url)
                        if p:
                            saved_rel.append(os.path.relpath(p, BASE_DIR).replace("\\", "/"))

                first_rel = saved_rel[0] if saved_rel else ""
                folder_rel = os.path.relpath(img_folder, BASE_DIR).replace("\\", "/")

                ws.append([
                    row, product_no, title, price, sale_price, shipping_fee,
                    review_count, rating, href, product_id,
                    folder_rel, first_rel, "; ".join(saved_rel), len(saved_rel)
                ])

                print(f"✅ {row}: {title} | 정가:{price} | 할인가:{sale_price} | 배송:{shipping_fee} | 리뷰:{review_count} | 평점:{rating} | imgs:{len(saved_rel)}")

                row += 1
                autosave_counter += 1
                if autosave_counter >= AUTOSAVE_EVERY:
                    safe_save_workbook(wb, EXCEL_PATH)
                    save_index(market_id, prod_map, last_no)
                    autosave_counter = 0

            except KeyboardInterrupt:
                print("\n🟥 사용자 중단 — 현재까지 진행분 저장합니다…")
                safe_save_workbook(wb, EXCEL_PATH)
                save_index(market_id, prod_map, last_no)
                raise
            except Exception as e:
                print("[SKIP]", e)
                continue
    finally:
        safe_save_workbook(wb, EXCEL_PATH)
        save_index(market_id, prod_map, last_no)
        print(f"📁 중간 저장 완료: {EXCEL_PATH}")

# ──────────────────────────────────────────────────────────────
# 인덱스
# ──────────────────────────────────────────────────────────────
def load_index(market_id: str) -> Tuple[Dict[str, str], int]:
    path = os.path.join(BASE_DIR, f"product_index_{market_id}.json")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("map", {}), int(data.get("last_no", 0))
    return {}, 0

def save_index(market_id: str, mapping: Dict[str, str], last_no: int):
    path = os.path.join(BASE_DIR, f"product_index_{market_id}.json")
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump({"map": mapping, "last_no": last_no}, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def assign_product_no(prod_id: str, mapping: Dict[str, str], last_no: int) -> Tuple[str, int]:
    if prod_id in mapping:
        return mapping[prod_id], last_no
    new_no = last_no + 1
    prod_no = f"P{new_no:06d}"
    mapping[prod_id] = prod_no
    return prod_no, new_no

# ──────────────────────────────────────────────────────────────
# 실행
# ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    driver, wait = make_driver()
    try:
        for url in MARKETS:
            collect_one_market(driver, wait, url)
        safe_save_workbook(wb, EXCEL_PATH)
        print("\n💾 엑셀 저장 완료 (정상 종료).")
    except KeyboardInterrupt:
        print("\n🟥 메인: 사용자 중단 감지 — 저장 후 종료합니다.")
        safe_save_workbook(wb, EXCEL_PATH)
    except Exception as e:
        print(f"\n🟠 메인 예외: {e} — 저장 후 종료합니다.")
        safe_save_workbook(wb, EXCEL_PATH)
    finally:
        try:
            safe_save_workbook(wb, EXCEL_PATH)
        except Exception as e:
            print(f"[SAVE ERROR] 최종 저장 실패: {e}")
        try:
            driver.quit()
        except:
            pass
        print("\n✅ 전체 크롤링 종료.")
